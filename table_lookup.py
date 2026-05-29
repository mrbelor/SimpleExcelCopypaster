from pathlib import Path
import datetime
import numbers
import pandas as pd


def _xl_cell_to_str(v):
    # openpyxl возвращает Python-объекты (datetime, int, float, str, None) —
    # конвертируем каждый тип в строку явно, без промежуточного str()
    if isinstance(v, datetime.datetime):
        if v.time() == datetime.time():
            return v.strftime("%Y-%m-%d")
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if v is None:
        return ""
    if isinstance(v, numbers.Integral):
        return str(int(v))
    if isinstance(v, numbers.Real):
        f = float(v)
        if pd.isna(f):
            return ""
        return str(int(f)) if f == int(f) else str(f)
    s = str(v)
    return "" if s == "nan" else s


def _load_as_strings(path, sheet_name=None):
    # всё читаем строками — никакой магии с типами
    suffix = path.suffix.lower()

    if suffix == ".csv":
        kwargs = {"dtype": str, "keep_default_na": False}
        df = None
        for enc in ("utf-8-sig", "utf-8", "cp1251"):
            try:
                df = pd.read_csv(path, encoding=enc, **kwargs)
                break
            except UnicodeDecodeError:
                continue
        if df is None:
            raise ValueError(f"не удалось прочитать CSV: {path}")

    elif suffix in (".xls", ".xlsx"):
        kw = {"keep_default_na": False}
        if sheet_name is not None:
            kw["sheet_name"] = sheet_name
        df = pd.read_excel(path, **kw)
        # применяем явную конвертацию до того, как pandas успевает звать str()
        df = df.apply(lambda col: col.map(_xl_cell_to_str))

    else:
        raise ValueError(f"формат '{suffix}' не поддерживается. используй .csv / .xls / .xlsx")

    # убираем пробелы из названий столбцов — в экселе они встречаются постоянно
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _find_output_col(df):
    # ищем первый безымянный столбец после последнего именованного
    # pandas называет пустые заголовки "Unnamed: N"
    cols = list(df.columns)
    last_named_idx = max(
        (i for i, c in enumerate(cols) if not c.startswith("Unnamed:")),
        default=-1,
    )
    target_idx = last_named_idx + 1
    if target_idx < len(cols):
        return cols[target_idx]
    # безымянных нет — вернём имя для нового столбца, который создаст _save
    return f"Unnamed: {len(cols)}"


class _Source:
    def __init__(self, path, sheet_name, key_col, val_col, template):
        self.path = Path(path)
        self.sheet_name = sheet_name
        self.key_col = key_col
        self.val_col = val_col
        self.template = template  # например "дата установки: {value}"
        self._df = None

    def load(self):
        self._df = _load_as_strings(self.path, self.sheet_name)
        sheet_info = f" / лист '{self.sheet_name}'" if self.sheet_name else ""
        cols = list(self._df.columns)
        for col, role in ((self.key_col, "key_column"), (self.val_col, "value_column")):
            if col not in cols:
                raise ValueError(
                    f"столбец '{col}' ({role}) не найден в {self.path.name}{sheet_info}. "
                    f"Доступные столбцы: {cols}"
                )

    def lookup(self, key):
        col_values = self._df[self.key_col].astype(str).str.strip()
        matches = self._df[col_values == key]
        if matches.empty:
            return None
        raw = str(matches.iloc[0][self.val_col]).strip()
        if raw == "nan":
            raw = ""
        if self.template:
            return self.template.format(value=raw)
        return raw


class TableLookup:
    def __init__(self, path, key_column, output_column=None, sheet_name=None):
        self.path = Path(path)
        self.key_column = key_column
        self.output_column = output_column  # None — найдём первый безымянный сами
        self.sheet_name = sheet_name
        self._sources = []

    def addSourceTable(self, path, key_column, value_column, sheet_name=None, template=None):
        self._sources.append(_Source(
            path=path,
            sheet_name=sheet_name,
            key_col=key_column,
            val_col=value_column,
            template=template,
        ))

    def run(self):
        df = _load_as_strings(self.path, self.sheet_name)

        for source in self._sources:
            source.load()
            sheet_info = f" / лист '{source.sheet_name}'" if source.sheet_name else ""
            print(f"  загружен: {source.path.name}{sheet_info}")

        # собираем результаты отдельно — не трогаем df до сохранения
        results = {}  # pandas-индекс → значение для записи
        for idx in df.index:
            key = str(df.at[idx, self.key_column]).strip()
            if not key or key == "nan":
                continue
            for source in self._sources:
                result = source.lookup(key)
                if result is not None:
                    results[idx] = result
                    break

        self._save(df, results)
        print(f"готово. обновлено строк: {len(results)} → {self.path}")

    def _save(self, df, results):
        suffix = self.path.suffix.lower()

        if suffix == ".csv":
            out_col = self.output_column or _find_output_col(df)
            if out_col not in df.columns:
                df[out_col] = ""
            for idx, value in results.items():
                df.at[idx, out_col] = value
            df.to_csv(self.path, index=False, encoding="utf-8-sig")

        elif suffix == ".xlsx":
            # xlsx: пишем только нужные ячейки — форматирование файла не трогаем
            self._save_xlsx_cells(results)

        elif suffix == ".xls":
            raise ValueError(
                ".xls — только чтение (старый формат). "
                "пересохрани главную таблицу как .xlsx или .csv"
            )
        else:
            raise ValueError(f"не умею писать в '{suffix}'. используй .csv или .xlsx")

    def _save_xlsx_cells(self, results):
        from openpyxl import load_workbook

        wb = load_workbook(self.path)
        ws_name = self.sheet_name if self.sheet_name is not None else wb.sheetnames[0]
        ws = wb[ws_name]

        # строим карту заголовок → номер столбца (1-based) по первой строке листа
        named_cols = {
            str(cell.value or "").strip(): cell.column
            for cell in ws[1]
            if str(cell.value or "").strip()
        }

        if self.key_column not in named_cols:
            raise ValueError(
                f"ключевой столбец '{self.key_column}' не найден в листе '{ws_name}'"
            )

        # определяем столбец для записи результата
        if self.output_column is not None:
            if self.output_column in named_cols:
                out_col_idx = named_cols[self.output_column]
            else:
                out_col_idx = ws.max_column + 1
                ws.cell(row=1, column=out_col_idx, value=self.output_column)
        else:
            # первый пустой столбец после последнего именованного
            last_named_col = max(named_cols.values(), default=0)
            out_col_idx = last_named_col + 1

        # пишем только изменённые ячейки; pandas-индекс 0 → Excel-строка 2
        for pandas_idx, value in results.items():
            ws.cell(row=pandas_idx + 2, column=out_col_idx, value=value)

        wb.save(self.path)
