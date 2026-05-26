import os
import sys
import shutil
import traceback
import pandas as pd
from pathlib import Path

# запускаться можно из любого места — корень проекта всегда shih/
PROJECT_ROOT = Path(__file__).parent.parent
os.chdir(PROJECT_ROOT)
sys.path.insert(0, str(PROJECT_ROOT))

from table_lookup import TableLookup

PASS = "  ✓"
FAIL = "  ✗"

results = []

def run(name, fn):
    try:
        fn()
        print(f"{PASS} {name}")
        results.append((name, True, None))
    except Exception as e:
        print(f"{FAIL} {name}")
        print(f"     {type(e).__name__}: {e}")
        traceback.print_exc()
        results.append((name, False, e))

def copy(src, dst):
    shutil.copy2(src, dst)
    return Path(dst)


# синтетические данные для тестов с реальными совпадениями
MATCHING_KEY = "999888777"

def make_main_csv(path, output_col=None):
    rows = [
        {"Договор": MATCHING_KEY, "Имя": "Тест", "Комментарий": ""},
        {"Договор": "000000000", "Имя": "Нет совпадения", "Комментарий": ""},
    ]
    df = pd.DataFrame(rows)
    if output_col is None:
        df[""] = ""
    df.to_csv(path, index=False, encoding="utf-8-sig")

def make_source_csv(path, key_col="ЛС", val_col="Результат"):
    df = pd.DataFrame([
        {key_col: MATCHING_KEY, val_col: "НАЙДЕНО"},
        {key_col: "111", val_col: "другое"},
    ])
    df.to_csv(path, index=False, encoding="utf-8-sig")

def make_main_xlsx(path, sheet="Лист1", output_col=None):
    rows = [
        {"Договор": MATCHING_KEY, "Имя": "Тест", "Комментарий": ""},
        {"Договор": "000000000", "Имя": "Нет совпадения", "Комментарий": ""},
    ]
    df = pd.DataFrame(rows)
    if output_col is None:
        df[""] = ""
    df.to_excel(path, sheet_name=sheet, index=False)

def make_source_xlsx(path, sheet="Данные", key_col="ЛС", val_col="Результат"):
    df = pd.DataFrame([
        {key_col: MATCHING_KEY, val_col: "НАЙДЕНО_XLSX"},
        {key_col: "111", val_col: "другое"},
    ])
    df.to_excel(path, sheet_name=sheet, index=False)


print("\n── реальные файлы: только чтение и загрузка ──\n")

def test_load_real_xlsx_main():
    from table_lookup import _load_as_strings
    df = _load_as_strings(Path("tests/sample_orders.xlsx"), sheet_name="Заказ")
    assert "Договор контрагента" in df.columns
    assert df["Договор контрагента"].iloc[0] == "40378028"

run("загрузка xlsx главной таблицы (лист 'Заказ')", test_load_real_xlsx_main)

def test_load_real_source_xlsx_выполненные():
    from table_lookup import _load_as_strings
    df = _load_as_strings(Path("tests/sample_completed.xlsx"), sheet_name="ФЛ")
    assert "Лицевой счет" in df.columns
    assert isinstance(df["Лицевой счет"].iloc[0], str)

run("загрузка xlsx реестра выполненных (Лицевой счет — строка)", test_load_real_source_xlsx_выполненные)

def test_load_real_source_xlsx_невыполненные_два_листа():
    from table_lookup import _load_as_strings
    df_fl = _load_as_strings(Path("tests/sample_pending.xlsx"), sheet_name="ФЛ")
    df_uv = _load_as_strings(Path("tests/sample_pending.xlsx"), sheet_name="УВЕД.")
    assert "ЛС" in df_fl.columns
    assert "ЛС" in df_uv.columns
    assert "Столбец15" in df_uv.columns

run("загрузка xlsx невыполненных (два листа: ФЛ и УВЕД.)", test_load_real_source_xlsx_невыполненные_два_листа)

def test_load_real_csv_source():
    from table_lookup import _load_as_strings
    df = _load_as_strings(Path("tests/sample_completed.csv"))
    assert "Лицевой счет" in df.columns
    assert isinstance(df["Лицевой счет"].iloc[0], str)

run("загрузка csv реестра выполненных (Лицевой счет — строка)", test_load_real_csv_source)

def test_key_never_changes_xlsx():
    tmp = copy("tests/sample_orders.xlsx", "/tmp/test_key_integrity.xlsx")
    before = pd.read_excel(tmp, sheet_name="Заказ", dtype=str, keep_default_na=False)["Договор контрагента"].tolist()

    lookup = TableLookup(path=tmp, key_column="Договор контрагента", sheet_name="Заказ")
    lookup.addSourceTable(
        path="tests/sample_completed.xlsx",
        sheet_name="ФЛ",
        key_column="Лицевой счет",
        value_column="ДАТА Установки ПУ",
    )
    lookup.run()

    after = pd.read_excel(tmp, sheet_name="Заказ", dtype=str, keep_default_na=False)["Договор контрагента"].tolist()
    assert before == after, f"ключ изменился! было {before}, стало {after}"

run("ключевое поле не изменяется после run() (xlsx)", test_key_never_changes_xlsx)


print("\n── синтетические данные: csv → csv ──\n")

def test_csv_to_csv_match():
    make_main_csv("/tmp/main_test.csv")
    make_source_csv("/tmp/src_test.csv")

    lookup = TableLookup(path="/tmp/main_test.csv", key_column="Договор")
    lookup.addSourceTable(path="/tmp/src_test.csv", key_column="ЛС", value_column="Результат")
    lookup.run()

    df = pd.read_csv("/tmp/main_test.csv", dtype=str, keep_default_na=False)
    unnamed = [c for c in df.columns if c.startswith("Unnamed:") or c == ""]
    assert unnamed, "нет безымянного столбца"
    out_col = unnamed[0]
    assert df[df["Договор"] == MATCHING_KEY][out_col].iloc[0] == "НАЙДЕНО"
    assert df[df["Договор"] == "000000000"][out_col].iloc[0] == ""

run("csv → csv: совпадение записывается, несовпадение пустое", test_csv_to_csv_match)

def test_csv_to_csv_template():
    make_main_csv("/tmp/main_tmpl.csv")
    make_source_csv("/tmp/src_tmpl.csv")

    lookup = TableLookup(path="/tmp/main_tmpl.csv", key_column="Договор")
    lookup.addSourceTable(
        path="/tmp/src_tmpl.csv",
        key_column="ЛС",
        value_column="Результат",
        template="ID заявки: {value}",
    )
    lookup.run()

    df = pd.read_csv("/tmp/main_tmpl.csv", dtype=str, keep_default_na=False)
    unnamed = [c for c in df.columns if c.startswith("Unnamed:") or c == ""]
    val = df[df["Договор"] == MATCHING_KEY][unnamed[0]].iloc[0]
    assert val == "ID заявки: НАЙДЕНО", f"получили: {repr(val)}"

run("csv → csv: шаблон template оборачивает значение", test_csv_to_csv_template)

def test_csv_explicit_output_column():
    make_main_csv("/tmp/main_expl.csv", output_col="explicit")
    make_source_csv("/tmp/src_expl.csv")

    lookup = TableLookup(path="/tmp/main_expl.csv", key_column="Договор", output_column="Комментарий")
    lookup.addSourceTable(path="/tmp/src_expl.csv", key_column="ЛС", value_column="Результат")
    lookup.run()

    df = pd.read_csv("/tmp/main_expl.csv", dtype=str, keep_default_na=False)
    assert df[df["Договор"] == MATCHING_KEY]["Комментарий"].iloc[0] == "НАЙДЕНО"

run("csv: явный output_column записывается в нужный столбец", test_csv_explicit_output_column)

def test_csv_multiple_sources_fallback():
    make_main_csv("/tmp/main_fallback.csv")
    pd.DataFrame([{"ЛС": "никогда", "Результат": "х"}]).to_csv("/tmp/src_empty.csv", index=False, encoding="utf-8-sig")
    make_source_csv("/tmp/src_second.csv")

    lookup = TableLookup(path="/tmp/main_fallback.csv", key_column="Договор")
    lookup.addSourceTable(path="/tmp/src_empty.csv", key_column="ЛС", value_column="Результат")
    lookup.addSourceTable(path="/tmp/src_second.csv", key_column="ЛС", value_column="Результат")
    lookup.run()

    df = pd.read_csv("/tmp/main_fallback.csv", dtype=str, keep_default_na=False)
    unnamed = [c for c in df.columns if c.startswith("Unnamed:") or c == ""]
    val = df[df["Договор"] == MATCHING_KEY][unnamed[0]].iloc[0]
    assert val == "НАЙДЕНО", f"ожидали НАЙДЕНО, получили: {repr(val)}"

run("csv: fallback — если первый источник не нашёл, идём во второй", test_csv_multiple_sources_fallback)

def test_csv_first_match_wins():
    make_main_csv("/tmp/main_firstwin.csv")
    pd.DataFrame([{"ЛС": MATCHING_KEY, "Результат": "ПЕРВЫЙ"}]).to_csv("/tmp/src_first.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame([{"ЛС": MATCHING_KEY, "Результат": "ВТОРОЙ"}]).to_csv("/tmp/src_second2.csv", index=False, encoding="utf-8-sig")

    lookup = TableLookup(path="/tmp/main_firstwin.csv", key_column="Договор")
    lookup.addSourceTable(path="/tmp/src_first.csv", key_column="ЛС", value_column="Результат")
    lookup.addSourceTable(path="/tmp/src_second2.csv", key_column="ЛС", value_column="Результат")
    lookup.run()

    df = pd.read_csv("/tmp/main_firstwin.csv", dtype=str, keep_default_na=False)
    unnamed = [c for c in df.columns if c.startswith("Unnamed:") or c == ""]
    val = df[df["Договор"] == MATCHING_KEY][unnamed[0]].iloc[0]
    assert val == "ПЕРВЫЙ", f"ожидали ПЕРВЫЙ, получили: {repr(val)}"

run("csv: первый источник с совпадением побеждает, второй игнорируется", test_csv_first_match_wins)


print("\n── синтетические данные: xlsx → xlsx ──\n")

def test_xlsx_to_xlsx_match():
    make_main_xlsx("/tmp/main_test.xlsx", sheet="Лист1")
    make_source_xlsx("/tmp/src_test.xlsx", sheet="Данные")

    lookup = TableLookup(path="/tmp/main_test.xlsx", key_column="Договор", sheet_name="Лист1")
    lookup.addSourceTable(path="/tmp/src_test.xlsx", sheet_name="Данные", key_column="ЛС", value_column="Результат")
    lookup.run()

    df = pd.read_excel("/tmp/main_test.xlsx", sheet_name="Лист1", dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    unnamed = [c for c in df.columns if c.startswith("Unnamed:") or c == ""]
    assert unnamed, f"нет безымянного столбца. колонки: {list(df.columns)}"
    val = df[df["Договор"] == MATCHING_KEY][unnamed[0]].iloc[0]
    assert val == "НАЙДЕНО_XLSX", f"получили: {repr(val)}"

run("xlsx → xlsx: совпадение записывается (с указанием листа)", test_xlsx_to_xlsx_match)

def test_xlsx_preserves_other_sheets():
    make_main_xlsx("/tmp/main_sheets.xlsx", sheet="Основной")
    with pd.ExcelWriter("/tmp/main_sheets.xlsx", engine="openpyxl", mode="a") as w:
        pd.DataFrame([{"x": 1}]).to_excel(w, sheet_name="НеТрогай", index=False)

    make_source_csv("/tmp/src_sheets.csv")

    lookup = TableLookup(path="/tmp/main_sheets.xlsx", key_column="Договор", sheet_name="Основной")
    lookup.addSourceTable(path="/tmp/src_sheets.csv", key_column="ЛС", value_column="Результат")
    lookup.run()

    xl = pd.ExcelFile("/tmp/main_sheets.xlsx")
    assert "НеТрогай" in xl.sheet_names, f"второй лист пропал! листы: {xl.sheet_names}"

run("xlsx: остальные листы не удаляются после run()", test_xlsx_preserves_other_sheets)

def test_xlsx_cell_formatting_preserved():
    # run() не должен сбрасывать форматирование — пишем только нужные ячейки
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    ws.append(["Договор", "Имя", ""])
    # красим заголовок — это форматирование должно выжить после run()
    red_fill = PatternFill(fill_type="solid", fgColor="FF0000")
    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = red_fill
        cell.font = bold_font
    ws.append([MATCHING_KEY, "Тест", ""])
    ws.append(["000000000", "Нет", ""])
    wb.save("/tmp/main_fmt.xlsx")

    make_source_csv("/tmp/src_fmt.csv")
    lookup = TableLookup(path="/tmp/main_fmt.xlsx", key_column="Договор", sheet_name="Лист1")
    lookup.addSourceTable(path="/tmp/src_fmt.csv", key_column="ЛС", value_column="Результат")
    lookup.run()

    wb2 = Workbook()
    from openpyxl import load_workbook
    wb2 = load_workbook("/tmp/main_fmt.xlsx")
    ws2 = wb2["Лист1"]
    assert ws2["A1"].fill.fill_type == "solid", "заливка заголовка сброшена"
    assert ws2["A1"].fill.fgColor.rgb[-6:] == "FF0000", "цвет заливки изменился"
    assert ws2["A1"].font.bold is True, "жирный шрифт заголовка сброшен"

run("xlsx: форматирование ячеек не сбрасывается после run()", test_xlsx_cell_formatting_preserved)

def test_xlsx_no_sheet_name_writes_to_first_sheet():
    make_main_xlsx("/tmp/main_nosheet.xlsx", sheet="Первый")
    make_source_csv("/tmp/src_nosheet.csv")

    lookup = TableLookup(path="/tmp/main_nosheet.xlsx", key_column="Договор", sheet_name=None)
    lookup.addSourceTable(path="/tmp/src_nosheet.csv", key_column="ЛС", value_column="Результат")
    lookup.run()

    df = pd.read_excel("/tmp/main_nosheet.xlsx", sheet_name="Первый", dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    unnamed = [c for c in df.columns if c.startswith("Unnamed:") or c == ""]
    assert unnamed, f"нет безымянного столбца в листе 'Первый'. колонки: {list(df.columns)}"
    val = df[df["Договор"] == MATCHING_KEY][unnamed[0]].iloc[0]
    assert val == "НАЙДЕНО", f"данные не записались в лист 'Первый', получили: {repr(val)}"

run("xlsx: sheet_name=None записывает в реальный первый лист, не в 'Sheet1'", test_xlsx_no_sheet_name_writes_to_first_sheet)


print("\n── смешанные форматы ──\n")

def test_xlsx_main_csv_source():
    make_main_xlsx("/tmp/main_mix1.xlsx", sheet="Лист1")
    make_source_csv("/tmp/src_mix1.csv")

    lookup = TableLookup(path="/tmp/main_mix1.xlsx", key_column="Договор", sheet_name="Лист1")
    lookup.addSourceTable(path="/tmp/src_mix1.csv", key_column="ЛС", value_column="Результат")
    lookup.run()

    df = pd.read_excel("/tmp/main_mix1.xlsx", sheet_name="Лист1", dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    unnamed = [c for c in df.columns if c.startswith("Unnamed:") or c == ""]
    val = df[df["Договор"] == MATCHING_KEY][unnamed[0]].iloc[0]
    assert val == "НАЙДЕНО", f"получили: {repr(val)}"

run("смешанный: xlsx главная + csv источник", test_xlsx_main_csv_source)

def test_csv_main_xlsx_source():
    make_main_csv("/tmp/main_mix2.csv")
    make_source_xlsx("/tmp/src_mix2.xlsx", sheet="Данные")

    lookup = TableLookup(path="/tmp/main_mix2.csv", key_column="Договор")
    lookup.addSourceTable(path="/tmp/src_mix2.xlsx", sheet_name="Данные", key_column="ЛС", value_column="Результат")
    lookup.run()

    df = pd.read_csv("/tmp/main_mix2.csv", dtype=str, keep_default_na=False)
    unnamed = [c for c in df.columns if c.startswith("Unnamed:") or c == ""]
    val = df[df["Договор"] == MATCHING_KEY][unnamed[0]].iloc[0]
    assert val == "НАЙДЕНО_XLSX", f"получили: {repr(val)}"

run("смешанный: csv главная + xlsx источник", test_csv_main_xlsx_source)

def test_xls_source_readable():
    from table_lookup import _load_as_strings
    df = _load_as_strings(Path("tests/sample_orders.xlsx"), sheet_name="Заказ")
    assert len(df) > 0
    assert isinstance(df["Договор контрагента"].iloc[0], str)

run("xls: читается как источник, ключ — строка", test_xls_source_readable)

def test_xls_main_raises_on_save():
    # создаём синтетический xls через xlwt или просто переименуем xlsx в xls для теста
    # xls как главная таблица — должна падать с понятной ошибкой при сохранении
    make_source_csv("/tmp/src_xls_err.csv")
    # копируем xlsx как xls в /tmp чтобы не трогать оригинал
    shutil.copy2("tests/sample_orders.xlsx", "/tmp/test_xls_main.xls")

    lookup = TableLookup(path="/tmp/test_xls_main.xls", key_column="Договор контрагента", sheet_name="Заказ")
    lookup.addSourceTable(path="/tmp/src_xls_err.csv", key_column="ЛС", value_column="Результат")

    try:
        lookup.run()
        raise AssertionError("должна была упасть с ValueError")
    except ValueError as e:
        assert ".xls" in str(e)

run("xls как главная: run() падает с понятной ошибкой (запись .xls невозможна)", test_xls_main_raises_on_save)


print("\n── граничные случаи ──\n")

def test_empty_key_rows_skipped():
    df = pd.DataFrame([
        {"Договор": "", "Результат_out": ""},
        {"Договор": MATCHING_KEY, "Результат_out": ""},
        {"Договор": "nan", "Результат_out": ""},
    ])
    df.to_csv("/tmp/main_empty_keys.csv", index=False, encoding="utf-8-sig")
    make_source_csv("/tmp/src_empty_keys.csv")

    lookup = TableLookup(path="/tmp/main_empty_keys.csv", key_column="Договор", output_column="Результат_out")
    lookup.addSourceTable(path="/tmp/src_empty_keys.csv", key_column="ЛС", value_column="Результат")
    lookup.run()

    result = pd.read_csv("/tmp/main_empty_keys.csv", dtype=str, keep_default_na=False)
    assert result[result["Договор"] == MATCHING_KEY]["Результат_out"].iloc[0] == "НАЙДЕНО"
    assert result[result["Договор"] == ""]["Результат_out"].iloc[0] == ""

run("граница: пустые ключи пропускаются, не роняют run()", test_empty_key_rows_skipped)

def test_whitespace_keys_match():
    df_main = pd.DataFrame([{"Договор": MATCHING_KEY, "out": ""}])
    df_main.to_csv("/tmp/main_ws.csv", index=False, encoding="utf-8-sig")
    df_src = pd.DataFrame([{"ЛС": f"  {MATCHING_KEY}  ", "Результат": "НАЙДЕНО_WS"}])
    df_src.to_csv("/tmp/src_ws.csv", index=False, encoding="utf-8-sig")

    lookup = TableLookup(path="/tmp/main_ws.csv", key_column="Договор", output_column="out")
    lookup.addSourceTable(path="/tmp/src_ws.csv", key_column="ЛС", value_column="Результат")
    lookup.run()

    result = pd.read_csv("/tmp/main_ws.csv", dtype=str, keep_default_na=False)
    assert result["out"].iloc[0] == "НАЙДЕНО_WS"

run("граница: пробелы вокруг ключа в источнике — всё равно совпадает", test_whitespace_keys_match)

def test_new_unnamed_col_created_when_none_exists():
    df = pd.DataFrame([{"A": MATCHING_KEY, "B": "доп"}])
    df.to_csv("/tmp/main_nonew.csv", index=False, encoding="utf-8-sig")
    make_source_csv("/tmp/src_nonew.csv")

    lookup = TableLookup(path="/tmp/main_nonew.csv", key_column="A")
    lookup.addSourceTable(path="/tmp/src_nonew.csv", key_column="ЛС", value_column="Результат")
    lookup.run()

    result = pd.read_csv("/tmp/main_nonew.csv", dtype=str, keep_default_na=False)
    assert len(result.columns) == 3, f"ожидали 3 столбца, получили {len(result.columns)}: {list(result.columns)}"
    last = result.columns[-1]
    assert result[last].iloc[0] == "НАЙДЕНО"

run("граница: нет безымянных столбцов — создаётся новый", test_new_unnamed_col_created_when_none_exists)

def test_excel_date_cells_read_as_plain_string():
    # Excel хранит даты как числа с типом date; openpyxl возвращает datetime-объекты.
    # Мы должны конвертировать их явно: дата → "YYYY-MM-DD", datetime → "YYYY-MM-DD HH:MM:SS"
    import datetime
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    ws.append(["Ключ", "Дата", "ДатаВремя"])
    ws.append([
        MATCHING_KEY,
        datetime.date(2026, 6, 5),
        datetime.datetime(2026, 6, 5, 14, 30, 0),
    ])
    wb.save("/tmp/main_dates.xlsx")

    from table_lookup import _load_as_strings
    df = _load_as_strings(Path("/tmp/main_dates.xlsx"), sheet_name="Лист1")

    date_val = df["Дата"].iloc[0]
    assert date_val == "2026-06-05", f"дата: '{date_val}' вместо '2026-06-05'"

    dt_val = df["ДатаВремя"].iloc[0]
    assert dt_val == "2026-06-05 14:30:00", f"дата+время: '{dt_val}' вместо '2026-06-05 14:30:00'"

run("граница: Excel-ячейки с типом date/datetime читаются корректно без артефактов", test_excel_date_cells_read_as_plain_string)


print("\n── итог ──\n")
passed = sum(1 for _, ok, _ in results if ok)
failed = sum(1 for _, ok, _ in results if not ok)
print(f"пройдено: {passed}/{len(results)}")
if failed:
    print(f"провалено: {failed}")
    for name, ok, err in results:
        if not ok:
            print(f"  ✗ {name}: {err}")
